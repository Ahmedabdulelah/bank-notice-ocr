from flask import Flask, render_template, request
import pytesseract
from PIL import Image, ImageOps, ImageFilter, ImageEnhance
import os
import re
from openpyxl import load_workbook 
from openpyxl.styles import Border, Side

app = Flask(__name__)

pytesseract.pytesseract.tesseract_cmd = r'C:\Users\ahmed\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'

UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/extract', methods=['POST'])
def extract_info():
    if 'image' not in request.files:
        return "No file part"
    file = request.files['image']
    if file.filename == '':
        return "No selected file"
    
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(file_path)
    
    image = preprocess_image(file_path)
    
    custom_config = r'--oem 3 --psm 3' 
    text = pytesseract.image_to_string(image, config=custom_config, lang='eng+ara') 
    
    extracted_data = extract_details(text)
    
    add_data_to_excel(extracted_data)

    return render_template('result.html', text=text, extracted_data=extracted_data)

def preprocess_image(image_path):
    image = Image.open(image_path)
    
    image = ImageOps.grayscale(image)
    
    enhancer = ImageEnhance.Contrast(image)
    image = enhancer.enhance(2.5) 
    image = image.filter(ImageFilter.SHARPEN)

    image = image.filter(ImageFilter.MedianFilter(size=3))

    image = image.point(lambda p: p > 180 and 255 or 0)

    return image

def clean_text(text):
    text = re.sub(r'[^\x00-\x7F]+', ' ', text)
    text = re.sub(r'\s+', ' ', text) 
    return text.strip()

def extract_details(text):
    data = {}
    
    text = clean_text(text)
    
    amount_match = re.search(r'\b(?:\d{1,3},)*\d{1,3}\.\d{2}\b', text)
    data['amount'] = amount_match.group(0) if amount_match else "Not found"
    
    datetime_match = re.search(r'\b(\d{2}-[A-Za-z]{3}-\d{4} \d{2}:\d{2}:\d{2})\b', text)
    if datetime_match:
        data['datetime'] = datetime_match.group(0)
    else:
        date_match = re.search(r'\b(\d{2}-[A-Za-z]{3}-\d{4})\b', text)
        data['date'] = date_match.group(0) if date_match else "Not found"
    
    return data

def add_data_to_excel(extracted_data):
    excel_file = r'C:\Users\ahmed\Desktop\اشعارت بنكك.xlsx'
    
    if not os.path.exists(excel_file):
        return "File does not exist!"
    
    wb = load_workbook(excel_file)
    ws = wb.active 

    row = 8  
    
    while ws.cell(row=row, column=5).value is not None: 
        row += 1  

    ws.cell(row=row, column=5, value=extracted_data.get('amount', 'Not found'))  # المبلغ بالجنية في E
    ws.cell(row=row, column=6, value=extracted_data.get('datetime', 'Not found'))  # التاريخ في F
    
    add_borders(ws, row, row, 5, 6)

    wb.save(excel_file)

def add_borders(ws, row_start, row_end, col_start, col_end):
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            ws.cell(row=row, column=col).border = thin_border

if __name__ == '__main__':
    app.run(debug=True)
